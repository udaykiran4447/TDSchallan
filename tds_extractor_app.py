import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TDS Challan Extractor", page_icon="📄", layout="wide")

st.markdown("""
<style>
    .main-header { font-size: 28px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }
    .sub-header { font-size: 14px; color: #6c757d; margin-bottom: 24px; }
    .metric-card { background: #f8f9fa; border-radius: 8px; padding: 16px; text-align: center; border: 1px solid #e9ecef; }
    .metric-value { font-size: 28px; font-weight: 700; color: #1a1a2e; }
    .metric-label { font-size: 12px; color: #6c757d; margin-top: 4px; }
    .stDataFrame { border-radius: 8px; }
    div[data-testid="stFileUploader"] { border: 2px dashed #dee2e6; border-radius: 12px; padding: 8px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">📄 TDS Challan PDF Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Upload ITNS 281 challan receipts — all data extracted into a single Excel sheet</div>', unsafe_allow_html=True)


def extract_value(text, label):
    patterns = [
        rf"{re.escape(label)}\s*[:\-]\s*(.+)",
        rf"{re.escape(label)}\s+(.+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def clean_amount(val):
    val = val.replace("₹", "").replace(",", "").strip()
    match = re.search(r"[\d]+(?:\.\d+)?", val)
    return float(match.group()) if match else 0.0


def extract_challan_data(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() + "\n"

    data = {}

    fields = {
        "TAN": "TAN",
        "Name": "Name",
        "Assessment Year": "Assessment Year",
        "Financial Year": "Financial Year",
        "Major Head": "Major Head",
        "Minor Head": "Minor Head",
        "Nature of Payment": "Nature of Payment",
        "CIN": "CIN",
        "Mode of Payment": "Mode of Payment",
        "Bank Name": "Bank Name",
        "Bank Reference Number": "Bank Reference Number",
        "Date of Deposit": "Date of Deposit",
        "BSR code": "BSR Code",
        "Challan No": "Challan No",
        "Tender Date": "Tender Date",
    }

    for label, key in fields.items():
        data[key] = extract_value(full_text, label)

    amount_match = re.search(r"Amount \(in Rs\.\)\s*[:\-]?\s*₹?\s*([\d,]+)", full_text)
    data["Amount (Rs.)"] = clean_amount(amount_match.group(1)) if amount_match else 0.0

    amount_words_match = re.search(r"Amount \(in words\)\s*[:\-]?\s*(.+)", full_text)
    data["Amount (in words)"] = amount_words_match.group(1).strip() if amount_words_match else ""

    breakup_fields = {
        "Tax": r"A\s+Tax\s+₹?\s*([\d,]+)",
        "Surcharge": r"B\s+Surcharge\s+₹?\s*([\d,]+)",
        "Cess": r"C\s+Cess\s+₹?\s*([\d,]+)",
        "Interest": r"D\s+Interest\s+₹?\s*([\d,]+)",
        "Penalty": r"E\s+Penalty\s+₹?\s*([\d,]+)",
        "Fee u/s 234E": r"F\s+Fee under section 234E\s+₹?\s*([\d,]+)",
        "Total": r"Total \(A\+B\+C\+D\+E\+F\)\s+₹?\s*([\d,]+)",
    }

    for key, pattern in breakup_fields.items():
        match = re.search(pattern, full_text)
        data[key] = clean_amount(match.group(1)) if match else 0.0

    itns_match = re.search(r"ITNS No\.\s*[:\-]?\s*(\d+)", full_text)
    data["ITNS No."] = itns_match.group(1).strip() if itns_match else ""

    return data


def create_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "TDS Challans"

    header_fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    sub_fill = PatternFill("solid", start_color="E8F4FD", end_color="E8F4FD")
    sub_font = Font(bold=True, name="Arial", size=9, color="1a1a2e")
    data_font = Font(name="Arial", size=9)
    alt_fill = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:T1")
    ws["A1"] = "TDS CHALLAN DETAILS — KAPSTON SERVICES LIMITED"
    ws["A1"].font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    main_headers = [
        "S.No", "ITNS No.", "TAN", "Name", "Assessment Year", "Financial Year",
        "Nature of Payment", "CIN", "Mode of Payment", "Bank Name",
        "Bank Ref. No.", "Date of Deposit", "BSR Code", "Challan No", "Tender Date",
        "Tax (Rs.)", "Surcharge (Rs.)", "Cess (Rs.)", "Interest (Rs.)",
        "Penalty (Rs.)", "Fee u/s 234E (Rs.)", "Total Amount (Rs.)"
    ]

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")

    for col_idx, header in enumerate(main_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.cell(row=3, column=col_idx).font = sub_font
        ws.cell(row=3, column=col_idx).fill = sub_fill
        ws.cell(row=3, column=col_idx).alignment = center
        ws.cell(row=3, column=col_idx).border = border

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    for i, rec in enumerate(records):
        row = i + 4
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        values = [
            i + 1,
            rec.get("ITNS No.", ""),
            rec.get("TAN", ""),
            rec.get("Name", ""),
            rec.get("Assessment Year", ""),
            rec.get("Financial Year", ""),
            rec.get("Nature of Payment", ""),
            rec.get("CIN", ""),
            rec.get("Mode of Payment", ""),
            rec.get("Bank Name", ""),
            rec.get("Bank Reference Number", ""),
            rec.get("Date of Deposit", ""),
            rec.get("BSR Code", ""),
            rec.get("Challan No", ""),
            rec.get("Tender Date", ""),
            rec.get("Tax", 0),
            rec.get("Surcharge", 0),
            rec.get("Cess", 0),
            rec.get("Interest", 0),
            rec.get("Penalty", 0),
            rec.get("Fee u/s 234E", 0),
            rec.get("Total", 0),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col_idx == 1 else left
            if col_idx >= 16:
                cell.number_format = '₹#,##0.00'

        ws.row_dimensions[row].height = 18

    total_row = len(records) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ws.cell(row=total_row, column=1).alignment = center
    ws.merge_cells(f"A{total_row}:O{total_row}")

    for col_idx in range(16, 23):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
        cell.number_format = '₹#,##0.00'
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[total_row].height = 20

    col_widths = [5, 8, 14, 28, 14, 12, 18, 26, 14, 14,
                  18, 14, 10, 10, 12, 14, 14, 10, 10, 10, 14, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A4"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


uploaded_files = st.file_uploader(
    "Upload challan PDF files",
    type=["pdf"],
    accept_multiple_files=True,
    help="Upload one or more ITNS 281 TDS challan PDF files"
)

if uploaded_files:
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(uploaded_files)}</div><div class="metric-label">Files Uploaded</div></div>', unsafe_allow_html=True)

    records = []
    errors = []

    with st.spinner("Extracting data from PDFs..."):
        for f in uploaded_files:
            try:
                data = extract_challan_data(f)
                data["_filename"] = f.name
                records.append(data)
            except Exception as e:
                errors.append((f.name, str(e)))

    if errors:
        for fname, err in errors:
            st.error(f"❌ {fname}: {err}")

    if records:
        total_amount = sum(r.get("Total", 0) for r in records)
        with col2:
            st.markdown(f'<div class="metric-card"><div class="metric-value">{len(records)}</div><div class="metric-label">Extracted Successfully</div></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="metric-card"><div class="metric-value">₹{total_amount:,.0f}</div><div class="metric-label">Total TDS Amount</div></div>', unsafe_allow_html=True)

        st.markdown("### 📊 Extracted Data Preview")

        preview_cols = [
            "Nature of Payment", "CIN", "Challan No", "Date of Deposit",
            "BSR Code", "Tax", "Surcharge", "Cess", "Interest", "Penalty", "Fee u/s 234E", "Total"
        ]
        df = pd.DataFrame(records)
        df.insert(0, "S.No", range(1, len(df) + 1))
        df["File"] = df["_filename"]

        display_cols = ["S.No", "File"] + [c for c in preview_cols if c in df.columns]
        st.dataframe(df[display_cols], use_container_width=True, hide_index=True)

        st.markdown("### 💾 Export to Excel")
        excel_data = create_excel(records)
        st.download_button(
            label="⬇️ Download Excel File",
            data=excel_data,
            file_name="TDS_Challans.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

else:
    st.info("👆 Upload one or more TDS challan PDF files to get started.")
    st.markdown("**Supported format:** ITNS 281 Challan Receipts from Income Tax Department")
